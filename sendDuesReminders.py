#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import smtplib
import sys

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
                   

lastCol = sheet.max_column

lastMonth = sheet.cell(row=1, column=lastCol).value

unpaidMembers = {}

for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=1).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

smtpdObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpdObj.ehlo()
smtpdObj.starttls()
smtpdObj.login('my_email_address@gmail.com', sys.argv[1])

for name, email in unpaidMembers.items():
    body = "Subject: %s dues unpaid.\nDear %s, \nRecords show that you have not \
    paid dues for %s. Please make this payment as soon as possible. Thank \
    you!'" % (lastMonth, name, lastMonth)
    print('Sending email to %s...' % (email))
    sendmailStatus = smtpdObj.sendmail('my_email_address@gmail.com', email, body)

    if sendmailStatus != {}:
        print('There was a problem sending mail to %s: %s' % (email,
        sendmailStatus))
smtpdObj.quit()

